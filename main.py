# 🧑‍💻 Project 3: Data Cleaning & Report Generator
# Author: Your Name
# Description:
# This script cleans messy sales data and generates a formatted Excel summary report.
# Deliverables:
#   1. sales_cleaned.xlsx — Cleaned dataset
#   2. sales_summary.xlsx — Summary report with formatting

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers

# -----------------------------
# Step 1: Load Raw Data
# -----------------------------
print("📥 Loading sales data...")
df = pd.read_csv("sales_raw.csv")

print("\n🔍 Raw Data Preview:")
print(df.head())

# -----------------------------
# Step 2: Clean Data
# -----------------------------
print("\n🧹 Cleaning data...")

# Remove duplicates
df = df.drop_duplicates()

# Fill missing Quantity with 1 (assuming missing quantities mean single item)
df['Quantity'] = df['Quantity'].fillna(1)

# Remove currency symbols and commas from Price and Total columns
for col in ['Price', 'Total']:
    df[col] = df[col].astype(str).replace(r'[^0-9.]', '', regex=True)
    df[col] = pd.to_numeric(df[col], errors='coerce')

# Handle missing or invalid numeric values
df['Price'] = df['Price'].fillna(0)
df['Total'] = df['Total'].fillna(df['Quantity'] * df['Price'])

# Recalculate total to ensure accuracy
df['Total'] = df['Quantity'] * df['Price']

# Clean column names (optional)
df.columns = df.columns.str.strip().str.title()

# -----------------------------
# Step 3: Save Cleaned Data
# -----------------------------
df.to_excel("sales_cleaned.xlsx", index=False)
print("✅ Cleaned data saved to 'sales_cleaned.xlsx'")

# -----------------------------
# Step 4: Generate Summary Report
# -----------------------------
print("\n📊 Generating summary report...")

# Total sales per product
sales_per_product = df.groupby('Product')['Total'].sum().reset_index()

# Calculate summary metrics
total_revenue = df['Total'].sum()
avg_order_value = df['Total'].mean()

# -----------------------------
# Step 5: Create Formatted Excel Report
# -----------------------------
wb = Workbook()
ws = wb.active
ws.title = "Sales Summary"

# Header row
headers = ["Product", "Total Sales ($)"]
ws.append(headers)

# Apply bold font to headers
for cell in ws[1]:
    cell.font = Font(bold=True)

# Add product-level data
for index, row in sales_per_product.iterrows():
    ws.append([row['Product'], row['Total']])

# Add empty row before summary
ws.append([])

# Add total and average rows
ws.append(["Total Revenue", total_revenue])
ws.append(["Average Order Value", avg_order_value])

# -----------------------------
# Step 6: Apply Formatting
# -----------------------------

# Apply currency format
for row in ws.iter_rows(min_row=2, min_col=2):
    for cell in row:
        if isinstance(cell.value, (int, float)):
            cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

# Format summary section (last two rows)
for row in ws.iter_rows(min_row=ws.max_row - 1, max_row=ws.max_row, min_col=1, max_col=2):
    for cell in row:
        cell.font = Font(bold=True, color="0000FF")  # Blue and bold

# Center align headers
for cell in ws[1]:
    cell.alignment = Alignment(horizontal="center")

# Adjust column width
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    ws.column_dimensions[column].width = max_length + 2

# -----------------------------
# Step 7: Save Summary Report
# -----------------------------
wb.save("sales_summary.xlsx")
print("✅ Summary report saved to 'sales_summary.xlsx'")

# -----------------------------
# Step 8: Completion Message
# -----------------------------
print("\n🎉 Data cleaning and report generation complete!")
print("Generated Files:")
print("  - sales_cleaned.xlsx")
print("  - sales_summary.xlsx")
